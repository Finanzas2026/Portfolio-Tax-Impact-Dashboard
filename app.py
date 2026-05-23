import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl, io, os, warnings
warnings.filterwarnings("ignore")

st.set_page_config(page_title="Tax Impact Dashboard", page_icon="📊", layout="wide",
                   initial_sidebar_state="collapsed")

st.markdown("""
<style>
.main .block-container{max-width:1400px;padding:1.5rem 2rem;overflow-x:auto;}
section[data-testid="stSidebar"]{display:none;}
.header-box{background:#4D93D9;padding:5px 32px;
    border-radius:12px;margin:0 auto 24px;width:80%;}
.header-title{color:white;font-size:26px;font-weight:900;letter-spacing:2px;}
.header-sub{color:#D0E8FF;font-size:13px;margin-top:4px;}
.zona-title{font-size:15px;font-weight:900;color:#0A2463;background:#DCE6F1;
    padding:1px 20px;border-radius:8px;margin:20px 0 14px;letter-spacing:1px;
    display:inline-block;min-width:460px;text-align:center;}
.kpi-box{background:#ffffff;border-radius:10px;padding:14px 18px;
    box-shadow:0 2px 8px rgba(0,0,0,.08);text-align:center;
    max-width:300px;width:300px;margin-right:10px;}
.kpi-lbl{font-size:10px;font-weight:700;color:#888;text-transform:uppercase;
    letter-spacing:.5px;margin-bottom:6px;}
.kpi-val{font-size:22px;font-weight:900;color:#0A2463;}
@media(max-width:768px){
    .main .block-container{padding:0.5rem !important;}
    .header-box{padding:14px 16px !important;}
    .header-title{font-size:16px !important;}
    .header-sub{font-size:11px !important;}
    .kpi-box{width:100% !important;max-width:100% !important;
        margin-right:0 !important;margin-bottom:8px;}
    .kpi-val{font-size:16px !important;}
    .zona-title{font-size:12px !important;padding:8px 12px !important;}
}
</style>""", unsafe_allow_html=True)

# ── DATOS ──────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def cargar():
    import requests
    RUTA_LOCAL = r"C:\Users\Arturo Aguilar\Documents\CLOSING\CASCO\Resumen_Inversiones_Casco.xlsx"
    if os.path.exists(RUTA_LOCAL):
        wb = openpyxl.load_workbook(RUTA_LOCAL, data_only=True)
    else:
        FILE_ID = st.secrets["FILE_ID"]
        r = requests.get(f"https://docs.google.com/spreadsheets/d/{FILE_ID}/export?format=xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["DATA"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append({"id": row[0], "zona": row[1], "proyecto": row[2],
                         "metrica": row[3], "actual": row[4], "sin_tx": row[5]})
    return pd.DataFrame(rows)

df_all = cargar()

def fmt_v(v, pct=False):
    if v is None or (isinstance(v, float) and pd.isna(v)): return "—"
    return f"{v*100:.2f}%" if pct else f"${v:,.0f}"

def es_pct_metrica(m):
    return m in ["CASH ON CASH", "IRR"]

# ── HEADER ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-box">
  <div class="header-title">PORTFOLIO — TAX IMPACT DASHBOARD</div>
  <div class="header-sub">Current Value vs Tax-Free Value · CASH ON CASH · IRR · FCF FROM FINANCING</div>
</div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN RESUMEN FCF FROM FINANCING
# ══════════════════════════════════════════════════════════════════════════════
def resumen_fcf(df):
    def get_fcf(zona):
        sub = df[(df["zona"] == zona) & (df["metrica"] == "FCF FROM FINANCING")]
        return sub["actual"].sum(), sub["sin_tx"].sum()

    a_ca, s_ca = get_fcf("CASCO ANTIGUO")
    a_sa, s_sa = get_fcf("SANTA ANA")
    a_tot = a_ca + a_sa
    s_tot = s_ca + s_sa

    st.markdown('<div style="text-align:center;"><div class="zona-title">PORTFOLIO SUMMARY — FCF FROM FINANCING</div></div>', unsafe_allow_html=True)
    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)

    def pct_var(base, diff):
        if not base or base == 0: return ""
        return f"{diff / base * 100:+.1f}%"

    def pct_of(part, total):
        if not total or total == 0: return ""
        return f"{part / total * 100:.1f}%"

    filas = [
        [("FCF With Tax — Casco Antiguo", f"${a_ca:,.0f}",  pct_of(a_ca,  a_tot)),
         ("FCF With Tax — Santa Ana",     f"${a_sa:,.0f}",  pct_of(a_sa,  a_tot)),
         ("FCF With Tax — Total",         f"${a_tot:,.0f}", pct_of(a_tot, a_tot))],
        [("FCF Tax-Free (Casco Law) — Casco Antiguo", f"${s_ca:,.0f}",  pct_of(s_ca,  s_tot)),
         ("FCF Tax-Free (Casco Law) — Santa Ana",     f"${s_sa:,.0f}",  pct_of(s_sa,  s_tot)),
         ("FCF Tax-Free (Casco Law) — Total",         f"${s_tot:,.0f}", pct_of(s_tot, s_tot))],
        [("FCF Variance — Casco Antiguo", f"${s_ca-a_ca:+,.0f}",     pct_var(a_ca,  s_ca-a_ca)),
         ("FCF Variance — Santa Ana",     f"${s_sa-a_sa:+,.0f}",     pct_var(a_sa,  s_sa-a_sa)),
         ("FCF Variance — Total",         f"${s_tot-a_tot:+,.0f}",   pct_var(a_tot, s_tot-a_tot))],
    ]
    for fila in filas:
        html = '<div style="display:flex;gap:10px;justify-content:center;margin-bottom:10px;">'
        for lbl, val, pct in fila:
            if pct:
                if pct.startswith("+"):
                    color = "#2ECC71"
                elif pct.startswith("-"):
                    color = "#E74C3C"
                else:
                    color = "#0070C0"
                pct_html = f'<div style="font-size:13px;font-weight:700;color:{color};margin-top:4px;">{pct}</div>'
            else:
                pct_html = ""
            html += (
                '<div class="kpi-box">'
                f'<div class="kpi-lbl">{lbl}</div>'
                f'<div class="kpi-val">{val}</div>'
                f'{pct_html}'
                '</div>'
            )
        html += '</div>'
        st.markdown(html, unsafe_allow_html=True)

    st.markdown("""
<div style="background:#F0F4FA;border-left:4px solid #0070C0;border-radius:6px;
     padding:12px 20px;margin:16px auto 8px auto;max-width:860px;font-size:13px;color:#333;line-height:1.8;">
  <b style="color:#0A2463;">FCF With Tax</b> — Free cash flow from financing applying the current tax burden (base scenario).<br>
  <b style="color:#0A2463;">FCF Tax-Free (Casco Law)</b> — Free cash flow under the tax exemption benefit of the Casco Antiguo Law.<br>
  <b style="color:#0A2463;">FCF Variance</b> — Absolute and relative difference between both scenarios; reflects the direct economic impact of the tax benefit.
</div>""", unsafe_allow_html=True)
    st.markdown("<hr style='border:none;border-top:2px solid #e0e0e0;width:80%;margin:24px auto;'>",
                unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SECCIÓN: render_zona
# ══════════════════════════════════════════════════════════════════════════════
def render_zona(zona, metrica_grafico, kpi_orden, default_proyecto):
    st.markdown(f'<div class="zona-title">{zona}</div>', unsafe_allow_html=True)

    df_z = df_all[df_all["zona"] == zona]
    proyectos = sorted(df_z["proyecto"].unique().tolist())

    # Dropdown ancho reducido (~4cm) con opción "Todos"
    opciones = ["All"] + proyectos
    default_idx = opciones.index(default_proyecto) if default_proyecto in opciones else 0
    col_sel, _ = st.columns([1, 4])
    with col_sel:
        proyecto_sel = st.selectbox("Project:", opciones, index=default_idx,
                                    key=f"sel_{zona}")

    df_proy = df_z.copy() if proyecto_sel == "All" else df_z[df_z["proyecto"] == proyecto_sel].copy()

    # ── 3 ETIQUETAS KPI ────────────────────────────────────────────────────────
    st.markdown("")
    cards_html = '<div style="display:flex;gap:10px;flex-wrap:wrap;justify-content:center;">'
    for m in kpi_orden:
        row_m = df_proy[df_proy["metrica"] == m]
        if row_m.empty:
            diff_val = "—"
            pct_html = ""
        else:
            actual = row_m["actual"].sum()
            sin_tx = row_m["sin_tx"].sum()
            diff   = sin_tx - actual
            diff_val = fmt_v(diff, pct=es_pct_metrica(m))
            if actual and actual != 0:
                pct_change = diff / abs(actual) * 100
                pct_str    = f"{pct_change:+.1f}%"
                color      = "#2ECC71" if pct_change >= 0 else "#E74C3C"
                pct_html   = f'<div style="font-size:13px;font-weight:700;color:{color};margin-top:4px;">{pct_str}</div>'
            else:
                pct_html = ""
        cards_html += (
            '<div class="kpi-box">'
            f'<div class="kpi-lbl">Variance {m}</div>'
            f'<div class="kpi-val">{diff_val}</div>'
            f'{pct_html}'
            '</div>'
        )
    cards_html += '</div>'
    st.markdown(cards_html, unsafe_allow_html=True)

    st.markdown("<div style='margin-top:100px;'></div>", unsafe_allow_html=True)

    # ── GRÁFICO: filtrado por selección ───────────────────────────────────────
    df_graf = df_proy[df_proy["metrica"] == metrica_grafico].copy()
    es_pct  = es_pct_metrica(metrica_grafico)

    fig = go.Figure()
    bar_width = 0.25 if proyecto_sel != "All" else 0.35
    fig.add_trace(go.Bar(
        name="With Tax", x=df_graf["proyecto"], y=df_graf["actual"],
        marker_color="#0A2463", width=bar_width,
        text=[f"{v*100:.1f}%" if es_pct else f"${v:,.0f}" for v in df_graf["actual"]],
        textposition="outside", textfont=dict(size=15)
    ))
    fig.add_trace(go.Bar(
        name="Tax-Free (Casco Law)", x=df_graf["proyecto"], y=df_graf["sin_tx"],
        marker_color="#4C9BE8", width=bar_width,
        text=[f"{v*100:.1f}%" if es_pct else f"${v:,.0f}" for v in df_graf["sin_tx"]],
        textposition="outside", textfont=dict(size=15)
    ))
    fig.update_layout(
        title=dict(text=f"{metrica_grafico} — {zona}", font=dict(size=13, color="#0A2463")),
        barmode="group", bargap=0.4, bargroupgap=0.08, plot_bgcolor="white",
        width=900, height=350,
        margin=dict(t=40, b=70, l=40, r=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        yaxis=dict(tickformat=".0%" if es_pct else "$,.0f", gridcolor="#F0F0F0"),
        xaxis=dict(tickangle=-30, tickfont=dict(size=12))
    )

    # Título y gráfico en la misma columna centrada
    nombre_filtro = proyecto_sel if proyecto_sel != "All" else "All Projects"
    _, col_c, _ = st.columns([0.3, 3, 0.3])
    with col_c:
        st.markdown(f"""
        <div style="background:#EFF3FA;border:1px solid #C0D0F0;border-radius:8px;
             padding:8px 16px;margin-bottom:8px;text-align:center;">
            <span style="font-size:19px;font-weight:700;color:#0A2463;">
                {nombre_filtro}
            </span>
        </div>""", unsafe_allow_html=True)
        fig.update_layout(width=None, height=350)
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("<div style='margin-top:30px;'></div>", unsafe_allow_html=True)

    # ── TABLA: TODOS los valores del proyecto seleccionado ────────────────────
    filas = []
    for _, r in df_proy.iterrows():
        m   = r["metrica"]
        pct = es_pct_metrica(m)
        a   = r["actual"]
        s   = r["sin_tx"]
        d   = s - a
        filas.append({
            "Description / Metric":  m,
            "With Tax":              fmt_v(a, pct),
            "Tax-Free (Casco Law)":  fmt_v(s, pct),
            "Difference":            (f"{d*100:+.2f}%" if pct else f"${d:+,.0f}"),
        })

    _, col_t, _ = st.columns([1, 2, 1])
    with col_t:
        st.markdown(f"**{proyecto_sel} — Full Detail**")
        st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

st.markdown("<div style='margin-top:48px;'></div>", unsafe_allow_html=True)
resumen_fcf(df_all)

# ══════════════════════════════════════════════════════════════════════════════
# CASCO ANTIGUO — layout de etiquetas (2 filas × 3 métricas)
# ══════════════════════════════════════════════════════════════════════════════
def render_zona_etiquetas(zona, metricas, default_proyecto):
    df_z = df_all[df_all["zona"] == zona]
    opciones = ["All"] + sorted(df_z["proyecto"].unique().tolist())
    default_idx = opciones.index(default_proyecto) if default_proyecto in opciones else 0
    _, col_sel, _ = st.columns([0.2, 0.8, 4])
    with col_sel:
        proyecto_sel = st.selectbox("Project:", opciones, index=default_idx,
                                    key=f"sel_etq_{zona}")

    st.markdown(f'<div style="text-align:center;"><div class="zona-title">{zona}</div></div>',
                unsafe_allow_html=True)

    df_p = df_z.copy() if proyecto_sel == "All" else df_z[df_z["proyecto"] == proyecto_sel].copy()

    def kpi_card(lbl, val, pct_str, pct_color):
        pct_html = (f'<div style="font-size:13px;font-weight:700;color:{pct_color};margin-top:4px;">{pct_str}</div>'
                    if pct_str else
                    '<div style="font-size:13px;margin-top:4px;visibility:hidden;">—</div>')
        return (
            '<div class="kpi-box">'
            f'<div class="kpi-lbl">{lbl}</div>'
            f'<div class="kpi-val">{val}</div>'
            f'{pct_html}'
            '</div>'
        )

    def render_fila(titulo, campo, color_titulo, show_pct=True):
        st.markdown(
            f'<div style="text-align:center;margin:18px 0 8px;">'
            f'<span style="font-size:13px;font-weight:700;color:{color_titulo};">{titulo}</span>'
            f'</div>', unsafe_allow_html=True)
        html = '<div style="display:flex;gap:10px;justify-content:center;flex-wrap:wrap;">'
        for m in metricas:
            row_m = df_p[df_p["metrica"] == m]
            if row_m.empty:
                html += kpi_card(m, "—", "", "")
                continue
            actual = row_m["actual"].sum()
            sin_tx = row_m["sin_tx"].sum()
            val    = row_m[campo].sum()
            base   = actual if campo == "sin_tx" else sin_tx
            diff   = val - base
            val_fmt = fmt_v(val, pct=es_pct_metrica(m))
            if show_pct and base and base != 0:
                if es_pct_metrica(m):
                    # IRR / CASH ON CASH: mostrar diferencia en puntos porcentuales
                    pp = diff * 100
                    pct_str  = f"{pp:+.2f} pp"
                    pct_color = "#2ECC71" if pp >= 0 else "#E74C3C"
                else:
                    # FCF: mostrar variación relativa
                    pct_change = diff / abs(base) * 100
                    pct_str    = f"{pct_change:+.1f}%"
                    pct_color  = "#2ECC71" if pct_change >= 0 else "#E74C3C"
            else:
                pct_str, pct_color = "", ""
            html += kpi_card(m, val_fmt, pct_str, pct_color)
        html += '</div>'
        st.markdown(html, unsafe_allow_html=True)

    st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
    render_fila("BASE SCENARIO — WITH TAX", "actual", "#0A2463", show_pct=False)
    st.markdown("<div style='margin-top:28px;'></div>", unsafe_allow_html=True)
    render_fila("TAX-FREE (CASCO LAW)", "sin_tx", "#0070C0", show_pct=True)
    st.markdown("""
<div style="background:#F0F4FA;border-left:4px solid #0070C0;border-radius:6px;
     padding:12px 20px;margin:20px auto 8px auto;max-width:860px;font-size:13px;color:#333;line-height:1.8;">
  <b style="color:#0A2463;">Base Scenario — With Tax</b> — Project metrics under the current tax burden without applying exemption benefits.<br>
  <b style="color:#0A2463;">Tax-Free (Casco Law)</b> — Project metrics applying the Casco Antiguo Law tax exemption. The percentage indicates the variance from the base scenario.
</div>""", unsafe_allow_html=True)
    st.markdown("<div style='margin-top:20px;'></div>", unsafe_allow_html=True)

render_zona_etiquetas(
    zona             = "CASCO ANTIGUO",
    metricas         = ["CASH ON CASH", "FCF FROM FINANCING", "IRR"],
    default_proyecto = "MANSION BALUARTE"
)

st.markdown("<hr style='border:none;border-top:2px solid #e0e0e0;width:80%;margin:32px auto;'>",
            unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SANTA ANA
# ══════════════════════════════════════════════════════════════════════════════
render_zona_etiquetas(
    zona             = "SANTA ANA",
    metricas         = ["CASH ON CASH", "FCF FROM FINANCING", "IRR"],
    default_proyecto = "CENTRA LINK"
)
